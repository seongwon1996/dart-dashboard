"""재무 데이터 가공 및 엑셀 워크북 생성 (pandas 없이 openpyxl만 사용)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# 비교에 사용할 표준 계정 (fnlttSinglAcnt.json 기준)
TARGET_ACCOUNTS = [
    "매출액",
    "영업이익",
    "당기순이익",
    "자산총계",
    "부채총계",
    "자본총계",
]

RATIO_NAMES = ["영업이익률(%)", "순이익률(%)", "부채비율(%)", "ROE(%)"]


def parse_amount(value: str | None) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def extract_key_accounts(rows: list[dict]) -> dict[str, float | None]:
    """DART fnlttSinglAcnt 응답 rows에서 표준 계정 금액만 뽑아낸다.

    "당기순이익(손실)"처럼 접미사가 붙는 계정명이 있어 접두 일치로 매칭한다.
    """
    result: dict[str, float | None] = {}
    for account in TARGET_ACCOUNTS:
        match = next((r for r in rows if r.get("account_nm", "").startswith(account)), None)
        result[account] = parse_amount(match["thstrm_amount"]) if match else None
    return result


def compute_ratios(accounts: dict[str, float | None]) -> dict[str, float | None]:
    revenue = accounts.get("매출액")
    op = accounts.get("영업이익")
    net = accounts.get("당기순이익")
    debt = accounts.get("부채총계")
    equity = accounts.get("자본총계")

    def pct(numer, denom):
        if numer is None or not denom:
            return None
        return round(numer / denom * 100, 2)

    return {
        "영업이익률(%)": pct(op, revenue),
        "순이익률(%)": pct(net, revenue),
        "부채비율(%)": pct(debt, equity),
        "ROE(%)": pct(net, equity),
    }


def _autosize_columns(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 3


def _bold_header(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_workbook(companies: list[dict], years: list[int]) -> Workbook:
    """companies: [{"name": str, "years": {year: {"accounts": dict, "ratios": dict, "fs_div": str}}}, ...]
    years: 내림차순 정렬된 연도 리스트
    """
    wb = Workbook()

    # 1) 요약비교(억원): 행=계정명, 열=회사_연도
    ws1 = wb.active
    ws1.title = "요약비교(억원)"
    header = ["계정명"] + [f"{c['name']}_{y}" for c in companies for y in years]
    ws1.append(header)
    for account in TARGET_ACCOUNTS:
        row = [account]
        for c in companies:
            for y in years:
                amount = c["years"].get(y, {}).get("accounts", {}).get(account)
                row.append(round(amount / 1e8, 1) if amount is not None else None)
        ws1.append(row)
    _bold_header(ws1)
    _autosize_columns(ws1)

    # 2) 재무비율: 행=회사, 열=연도_지표
    ws2 = wb.create_sheet("재무비율")
    header2 = ["회사"] + [f"{y}_{r}" for y in years for r in RATIO_NAMES]
    ws2.append(header2)
    for c in companies:
        row = [c["name"]]
        for y in years:
            ratios = c["years"].get(y, {}).get("ratios", {})
            for r in RATIO_NAMES:
                row.append(ratios.get(r))
        ws2.append(row)
    _bold_header(ws2)
    _autosize_columns(ws2)

    # 3) 원본데이터: long format
    ws3 = wb.create_sheet("원본데이터")
    ws3.append(["회사", "연도", "계정명", "금액", "재무제표구분"])
    for c in companies:
        for y in years:
            year_data = c["years"].get(y)
            if not year_data:
                continue
            fs_div = year_data.get("fs_div")
            for account, amount in year_data.get("accounts", {}).items():
                ws3.append([c["name"], y, account, amount, fs_div])
    _bold_header(ws3)
    _autosize_columns(ws3)

    return wb
